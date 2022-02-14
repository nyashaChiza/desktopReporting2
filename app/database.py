# -*- coding: utf-8 -*-
# app/database.py
import sys,sqlite3
from datetime import date

def createConnection(databaseName):
        conn = sqlite3.connect(databaseName)
        c = conn.cursor()
        _createTables(c)
        return c

def _createTables(con):
    today = date.today()
    con.execute("CREATE TABLE IF NOT EXISTS A_I( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,sin TEXT,gender TEXT,first_name TEXT,last_name TEXT, last_name_prefix TEXT,academic_title TEXT,filler1 TEXT,birth_date DATE, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS J_S( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,country_code TEXT,postal_code TEXT,city TEXT,c_o TEXT, filler2 TEXT,filler3 TEXT,filler4 TEXT,country_code_mail_address TEXT, postal_code_mail_address TEXT,city_mail_address TEXT,date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS T_AC( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,date_of_entry_into_group TEXT,date_of_leaving_group TEXT,reason_for_leaving_the_group TEXT,bank_code TEXT, bank_account_number TEXT,swift_bic TEXT,owner_of_bank_account TEXT,employee_with_potential DATE,indicator_for_executive TEXT,indicator_for_security_advisor TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS AD_AN( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,indicator_eligibility_stock_purchase_program TEXT,company_code TEXT,personal_number TEXT,plant_identifier_for_personal_number TEXT, filler5 TEXT,filler6 TEXT,transfer_date TEXT,filler7 TEXT,filler8 TEXT,local_cost_center TEXT, date_entry DATE DEFAULT today )")
    
    con.execute("CREATE TABLE IF NOT EXISTS AM_AW( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,employee_group TEXT,filler9 TEXT,department_abbreviation TEXT,hr_executive_level TEXT, employment_type TEXT,organizational_code TEXT,physical_work_location_code TEXT,internal_mail_code TEXT,level_of_business_allocation1 TEXT,level_of_business_allocation2 TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS AX_BG( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,mailing_language_of_employee TEXT,building TEXT,room_number TEXT,location_code TEXT, country_of_birth TEXT,city_of_birth TEXT,citizenship TEXT,filler10 TEXT,filler11 TEXT,street_and_house_number_of_home_address TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS BH_BQ( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,street_and_house_number_for_preferred_mailing_address TEXT,filler12 TEXT,job_code TEXT,plant2 TEXT, plant1 TEXT,organizational_code2 TEXT,smtp_email_address TEXT,position_entry_date TEXT,date_contract_end TEXT,hr_department TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS BR_CA( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,hr_representative TEXT,fax_number_hr_representative TEXT,management_level TEXT,work_phone_number_of_hr_representative TEXT, preferred_first_name TEXT,middle_initial TEXT,home_host_indicator TEXT,fulltime_parttime_equivalency TEXT,reports_to_kim TEXT,date_of_entry_into_company TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS CB_CW( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,filler12 TEXT,filler13 TEXT,filler14 TEXT, filler15 TEXT,filler16 TEXT,filler17 TEXT,filler18 TEXT,filler19 TEXT,filler20 TEXT,filler21 TEXT,filler22 TEXT, filler23 TEXT,filler24 TEXT,filler25 TEXT,filler26 TEXT,filler27 TEXT,filler28 TEXT,filler29 TEXT,filler30 TEXT,filler31 TEXT,filler32 TEXT,filler33 TEXT, date_entry DATE DEFAULT today )")                 

    con.execute("CREATE TABLE IF NOT EXISTS CX_DG( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,current_position_title TEXT,payment_type TEXT,plant_section TEXT,filler35 TEXT, indicator_for_disabled_person TEXT,social_security_number TEXT,state TEXT,internal_mail_code TEXT,state_for_preferred_mailing_address TEXT,phone_country_code TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS DH_DQ( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,phone_country_code TEXT,phone_extension_and_company TEXT,fax_country_code TEXT,fax_area_code TEXT, fax_extension_and_company TEXT,mobile_country_code TEXT,mobile_area_code TEXT,mobile_number TEXT,empl_id TEXT,filler36 TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS DR_EA( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,filler37 TEXT,bv_wechsel_ab TEXT,wohnhaft_bei TEXT,company_code_expat TEXT, currency TEXT,filler38 TEXT,year_of_goal_income TEXT,monthly_salary TEXT,delete_flag TEXT,filler39 TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS EB_EK( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,grade_band TEXT,shift_indicator TEXT,employee_category TEXT,department_name TEXT, status_employee TEXT,code_dl_tv TEXT,grade TEXT,nacos_cost_center TEXT,wage_group TEXT,year_of_wage_group TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS EL_EY( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,date_of_assignment_to_wage_group TEXT,filler40 TEXT,logon_id_of_employee TEXT,company_name_of_external_person TEXT, filler41 TEXT,filler42 TEXT,filler43 TEXT,filler44 TEXT,filler45 TEXT,filler46 TEXT,filler47 TEXT,filler48 TEXT,filler49 TEXT,filler50 TEXT, date_entry DATE DEFAULT today )")

    con.execute("CREATE TABLE IF NOT EXISTS EZ_FI( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,bv_wechsel_ab TEXT,filler51 TEXT,filler52 TEXT,type_of_timeframe_for_tax_data TEXT, starting_of_timeframe_for_tax_data TEXT,end_of_timeframe_for_tax_data TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS FJ_FS( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,starting_of_timeframe_for_tax_data TEXT,end_of_timeframe_for_tax_data TEXT, type_of_timeframe_for_tax_data TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS FT_GC( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,end_of_timeframe_for_tax_data TEXT,type_of_timeframe_for_tax_data TEXT,starting_of_timeframe_for_tax_data TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS GD_GM( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,type_of_timeframe_for_tax_data TEXT,starting_of_timeframe_for_tax_data TEXT,end_of_timeframe_for_tax_data TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS GN_GW( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,permanent_residence TEXT,filler54 TEXT,wage_and_salary_group TEXT,diversity TEXT, executive_bonus TEXT,number_of_monthly_base_salary TEXT,structure_code TEXT,dept_id TEXT,full_time_equivalent TEXT,bonus_payout TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS GX_HK( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,position_control_number TEXT,daimler_cost_center TEXT,last_name_passport TEXT,first_name_passport TEXT, executive_bonus TEXT,iban TEXT,local_job_profile TEXT,irwwh TEXT,buchungskreis TEXT,rabattkennzeichen TEXT,valid_by TEXT,widersprecherkennzeichen TEXT,ivv_contract TEXT,ivv_risk_taker TEXT,ivv_control_unit TEXT, date_entry DATE DEFAULT today )")
    con.close()


#-----------------------------------------------------------------------------------------------------------------------------------
def updateInfo(data):
    conn = sqlite3.connect('database.sqlite')
    con = conn.cursor()  
    try:
        sql =  ''' UPDATE A_I
                    SET first_name = ?,
                    last_name = ?,
                    birth_date = ?
                    WHERE kim = ?'''
        con.execute(sql, (data.get('fname'),data.get('lname'),data.get('bdate'),data.get('kim')))

        sql =  '''UPDATE T_AC
        SET date_of_entry_into_group = ?
                    WHERE kim = ?'''
        con.execute(sql, [data.get('date_of_entry'),data.get('kim')])

        sql =  ''' UPDATE AD_AN
        SET personal_number = ? 
                    WHERE kim = ?'''
        con.execute(sql, [data.get('pnumber'),data.get('kim')])

        sql =  ''' UPDATE AM_AW
        SET employee_group = ? 
                    WHERE kim = ?'''
        con.execute(sql, [data.get('employee_group'),data.get('kim')])

        sql =  ''' UPDATE AM_AW
        SET hr_executive_level = ?
                    WHERE kim = ?'''
        con.execute(sql, [data.get('hr_level'),data.get('kim')])

        sql =  ''' UPDATE BH_BQ
        SET smtp_email_address = ?
                    WHERE kim = ?'''
        con.execute(sql, [data.get('email'),data.get('kim')])

        sql =  ''' UPDATE BR_CA
        SET reports_to_kim = ?
                    WHERE kim = ?'''
        con.execute(sql, [data.get('reports'),data.get('kim')])

        sql =  ''' UPDATE CX_DG
        SET current_position_title = ?
                    WHERE kim = ?'''
        con.execute(sql, [data.get('current_position'),data.get('kim')])
        conn.commit()
        return 1
    except Exception as e:
        print(e)
        return 0
#-----------------------------------------------------------------------------------------------------------------------------------
def loadInfo():
    from openpyxl import load_workbook
    conn = sqlite3.connect('database.sqlite')
    con = conn.cursor()  
    file = 'test.xlsx'
    wb_obj = load_workbook(filename=file)
    ws = wb_obj.active
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        #try:
            print('loading...')
            con.execute('''INSERT INTO A_I (kim,sin,gender,first_name,last_name,last_name_prefix,academic_title,filler1,birth_date,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value ,row[8].value,date.today() ))
            con.execute('''INSERT INTO J_S (kim,country_code,postal_code,city,c_o,filler2,filler3,filler4,country_code_mail_address,postal_code_mail_address,city_mail_address,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[9].value,row[10].value,row[11].value,row[12].value,row[13].value,row[14].value,row[15].value,row[16].value,row[17].value,row[18].value,date.today() ))
            con.execute('''INSERT INTO T_AC (kim,date_of_entry_into_group,date_of_leaving_group,reason_for_leaving_the_group,bank_code,bank_account_number,swift_bic,owner_of_bank_account,employee_with_potential,indicator_for_executive,indicator_for_security_advisor,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[19].value,row[20].value,row[21].value,row[22].value,row[23].value,row[24].value,row[25].value,row[26].value,row[27].value,row[28].value,date.today() ))
            con.execute('''INSERT INTO AD_AN (kim,indicator_eligibility_stock_purchase_program,company_code,personal_number,plant_identifier_for_personal_number,filler5,filler6,transfer_date,filler7,filler8,local_cost_center,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[29].value,row[30].value,row[31].value,row[32].value,row[33].value,row[34].value,row[35].value,row[36].value,row[37].value,row[38].value,date.today() ))
            con.execute('''INSERT INTO AM_AW (kim,employee_group,filler9,department_abbreviation,hr_executive_level,employment_type,organizational_code,physical_work_location_code,internal_mail_code,level_of_business_allocation1,level_of_business_allocation2,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[39].value,row[40].value,row[41].value,row[42].value,row[43].value,row[44].value,row[45].value,row[46].value,row[47].value,row[48].value,date.today() ))
            
            con.execute('''INSERT INTO AX_BG (kim,mailing_language_of_employee,building,room_number,location_code,country_of_birth,city_of_birth,citizenship,filler10,filler11,street_and_house_number_of_home_address,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[49].value,row[50].value,row[51].value,row[52].value,row[53].value,row[54].value,row[55].value,row[56].value,row[57].value,row[58].value,date.today() ))
            con.execute('''INSERT INTO BH_BQ (kim,street_and_house_number_for_preferred_mailing_address,filler12,job_code,plant1,plant2,organizational_code2,smtp_email_address,position_entry_date,date_contract_end,hr_department,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[59].value,row[60].value,row[61].value,row[62].value,row[64].value,row[63].value,row[65].value,row[66].value,row[67].value,row[68].value,date.today() ))
            con.execute('''INSERT INTO BR_CA (kim,hr_representative,fax_number_hr_representative,management_level,work_phone_number_of_hr_representative,preferred_first_name,middle_initial,home_host_indicator,fulltime_parttime_equivalency,reports_to_kim,date_of_entry_into_company,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[69].value,row[70].value,row[71].value,row[72].value,row[73].value,row[74].value,row[75].value,row[76].value,row[77].value,row[78].value,date.today() ))
            con.execute('''INSERT INTO CB_CW (kim,filler12,filler13,filler14,filler15,filler16,filler17,filler18,filler19,filler20,filler21,filler22,filler23,filler24,filler25,filler26,filler27,filler28,filler29,filler30,filler31,filler32,filler33,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[79].value,row[80].value,row[81].value,row[82].value,row[83].value,row[84].value,row[85].value,row[86].value,row[87].value,row[88].value,row[89].value,row[90].value,row[91].value,row[92].value,row[93].value,row[94].value,row[95].value,row[96].value,row[97].value,row[98].value,row[99].value,row[100].value,date.today() ))
            con.execute('''INSERT INTO CX_DG (kim,current_position_title,payment_type,plant_section,filler35,indicator_for_disabled_person,social_security_number,state,state_for_preferred_mailing_address,phone_country_code,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[101].value,row[98].value,row[102].value,row[103].value,row[104].value,row[105].value,row[106].value,row[107].value,row[108].value,date.today() ))
            
            con.execute('''INSERT INTO DH_DQ (kim,phone_country_code,phone_extension_and_company,fax_country_code,fax_area_code,fax_extension_and_company,mobile_country_code,mobile_area_code,mobile_number,empl_id,filler36,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[110].value,row[111].value,row[112].value,row[113].value,row[114].value,row[115].value,row[116].value,row[117].value,row[118].value,row[119].value,date.today() ))
            con.execute('''INSERT INTO DR_EA (kim,filler37,bv_wechsel_ab,wohnhaft_bei,company_code_expat,currency,filler38,year_of_goal_income,monthly_salary,delete_flag,filler39,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EB_EK (kim,grade_band,shift_indicator,employee_category,department_name,status_employee,status_employee,code_dl_tv,grade,nacos_cost_center,wage_group,year_of_wage_group,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EL_EY (kim,date_of_assignment_to_wage_group,filler40,logon_id_of_employee,company_name_of_external_person,filler41,filler42,filler43,filler44,filler45,filler46,filler47,filler48,filler49,filler50,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EZ_FI (kim,bv_wechsel_ab,filler51,filler52,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,date_entry)VALUES(?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))

            con.execute('''INSERT INTO FJ_FS (kim,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,date_entry)VALUES(?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO FT_GC (kim,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,date_entry)VALUES(?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GD_GM (kim,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,date_entry)VALUES(?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GN_GW (kim,permanent_residence,wage_and_salary_group,diversity,executive_bonus,number_of_monthly_base_salary,structure_code,dept_id,full_time_equivalent,bonus_payout,filler54,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GX_HK (kim,executive_bonus,rabattkennzeichen,position_control_number,daimler_cost_center,last_name_passport,last_name_passport,last_name_passport,first_name_passport,iban,local_job_profile,irwwh,valid_by,buchungskreis,widersprecherkennzeichen,ivv_contract,ivv_risk_taker,ivv_control_unit,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            conn.commit()
            #return 1
       # except Exception as e: 
       #     print(e)
            #break
       #     return e

#-----------------------------------------------------------------------------------------------------------------------------------
def saveInfo(data):
    from openpyxl import load_workbook
    conn = sqlite3.connect('database.sqlite')
    con = conn.cursor()  
    file = 'test.xlsx'
    wb_obj = load_workbook(filename=file)
    ws = wb_obj.active
    for row in ws.iter_rows(min_row=10, min_col=1, max_row=11, max_col=ws.max_column):
        #try:
            con.execute('''INSERT INTO A_I (kim,sin,gender,first_name,last_name,last_name_prefix,academic_title,filler1,birth_date,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[1].value,data.get('gender'),data.get('fname'),data.get('lname'),row[5].value,row[6].value,row[7].value, data.get('bdate'),date.today() ))
            con.execute('''INSERT INTO J_S (kim,country_code,postal_code,city,c_o,filler2,filler3,filler4,country_code_mail_address,postal_code_mail_address,city_mail_address,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[9].value,row[10].value,row[11].value,row[12].value,row[13].value,row[14].value,row[15].value,row[16].value,row[17].value,row[18].value,date.today() ))
            con.execute('''INSERT INTO T_AC (kim,date_of_entry_into_group,date_of_leaving_group,reason_for_leaving_the_group,bank_code,bank_account_number,swift_bic,owner_of_bank_account,employee_with_potential,indicator_for_executive,indicator_for_security_advisor,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),data.get('date_of_entry'),row[20].value,row[21].value,row[22].value,row[23].value,row[24].value,row[25].value,row[26].value,row[27].value,row[28].value,date.today() ))
            con.execute('''INSERT INTO AD_AN (kim,indicator_eligibility_stock_purchase_program,company_code,personal_number,plant_identifier_for_personal_number,filler5,filler6,transfer_date,filler7,filler8,local_cost_center,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[29].value,row[30].value,data.get('personal_number'),row[32].value,row[33].value,row[34].value,row[35].value,row[36].value,row[37].value,row[38].value,date.today() ))
            con.execute('''INSERT INTO AM_AW (kim,employee_group,filler9,department_abbreviation,hr_executive_level,employment_type,organizational_code,physical_work_location_code,internal_mail_code,level_of_business_allocation1,level_of_business_allocation2,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),data.get('employee_group'),row[40].value,row[41].value,data.get('hr_level'),data.get('etype'),row[44].value,row[45].value,row[46].value,row[47].value,row[48].value,date.today() ))
            
            con.execute('''INSERT INTO AX_BG (kim,mailing_language_of_employee,building,room_number,location_code,country_of_birth,city_of_birth,citizenship,filler10,filler11,street_and_house_number_of_home_address,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[49].value,row[50].value,row[51].value,row[52].value,row[53].value,row[54].value,row[55].value,row[56].value,row[57].value,row[58].value,date.today() ))
            con.execute('''INSERT INTO BH_BQ (kim,street_and_house_number_for_preferred_mailing_address,filler12,job_code,plant1,plant2,organizational_code2,smtp_email_address,position_entry_date,date_contract_end,hr_department,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[59].value,row[60].value,row[61].value,row[62].value,row[64].value,row[63].value,data.get('temail'),row[66].value,row[67].value,row[68].value,date.today() ))
            con.execute('''INSERT INTO BR_CA (kim,hr_representative,fax_number_hr_representative,management_level,work_phone_number_of_hr_representative,preferred_first_name,middle_initial,home_host_indicator,fulltime_parttime_equivalency,reports_to_kim,date_of_entry_into_company,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[69].value,row[70].value,row[71].value,row[72].value,row[73].value,row[74].value,row[75].value,row[76].value,data.get('report'),row[78].value,date.today() ))
            con.execute('''INSERT INTO CB_CW (kim,filler12,filler13,filler14,filler15,filler16,filler17,filler18,filler19,filler20,filler21,filler22,filler23,filler24,filler25,filler26,filler27,filler28,filler29,filler30,filler31,filler32,filler33,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[79].value,row[80].value,row[81].value,row[82].value,row[83].value,row[84].value,row[85].value,row[86].value,row[87].value,row[88].value,row[89].value,row[90].value,row[91].value,row[92].value,row[93].value,row[94].value,row[95].value,row[96].value,row[97].value,row[98].value,row[99].value,row[100].value,date.today() ))
            con.execute('''INSERT INTO CX_DG (kim,current_position_title,payment_type,plant_section,filler35,indicator_for_disabled_person,social_security_number,state,state_for_preferred_mailing_address,phone_country_code,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),data.get('position'),row[98].value,row[102].value,row[103].value,row[104].value,row[105].value,row[106].value,row[107].value,row[108].value,date.today() ))
            
            con.execute('''INSERT INTO DH_DQ (kim,phone_country_code,phone_extension_and_company,fax_country_code,fax_area_code,fax_extension_and_company,mobile_country_code,mobile_area_code,mobile_number,empl_id,filler36,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[110].value,row[111].value,row[112].value,row[113].value,row[114].value,row[115].value,row[116].value,row[117].value,row[118].value,row[119].value,date.today() ))
            con.execute('''INSERT INTO DR_EA (kim,filler37,bv_wechsel_ab,wohnhaft_bei,company_code_expat,currency,filler38,year_of_goal_income,monthly_salary,delete_flag,filler39,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EB_EK (kim,grade_band,shift_indicator,employee_category,department_name,status_employee,status_employee,code_dl_tv,grade,nacos_cost_center,wage_group,year_of_wage_group,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EL_EY (kim,date_of_assignment_to_wage_group,filler40,logon_id_of_employee,company_name_of_external_person,filler41,filler42,filler43,filler44,filler45,filler46,filler47,filler48,filler49,filler50,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EZ_FI (kim,bv_wechsel_ab,filler51,filler52,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,date_entry)VALUES(?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))

            con.execute('''INSERT INTO FJ_FS (kim,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,date_entry)VALUES(?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO FT_GC (kim,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,date_entry)VALUES(?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GD_GM (kim,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,date_entry)VALUES(?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GN_GW (kim,permanent_residence,wage_and_salary_group,diversity,executive_bonus,number_of_monthly_base_salary,structure_code,dept_id,full_time_equivalent,bonus_payout,filler54,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GX_HK (kim,executive_bonus,rabattkennzeichen,position_control_number,daimler_cost_center,last_name_passport,last_name_passport,last_name_passport,first_name_passport,iban,local_job_profile,irwwh,valid_by,buchungskreis,widersprecherkennzeichen,ivv_contract,ivv_risk_taker,ivv_control_unit,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            conn.commit()
            return 1
       # except Exception as e: 
       #     print(e)
            #break
       #     return e
#---------------------------------------------------------------------------------------------------------------------



def find_employee(kim):
    conn = sqlite3.connect('database.sqlite')
    con = conn.cursor()
    con.execute("SELECT * FROM A_I WHERE kim=?", (str(kim),))
    rows = con.fetchall()

    if len(rows) > 0:
        return 1
    else:
        return 0
#---------------------------------------------------------------------------------------------------------------------



