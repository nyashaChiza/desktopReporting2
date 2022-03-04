# -*- coding: utf-8 -*-
# app/database.py
import sys,sqlite3
from datetime import date
from openpyxl import load_workbook

#-----------------------------------------------------------------------------------------------------------------------------------
def createConnection(databaseName):
        conn = sqlite3.connect(databaseName)
        c = conn.cursor()
        _createTables(c)
        return c
#-----------------------------------------------------------------------------------------------------------------------------------
def _createTables(con):

    con.execute("CREATE TABLE IF NOT EXISTS A_I( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,sin TEXT,gender TEXT,first_name TEXT,last_name TEXT, last_name_prefix TEXT,academic_title TEXT,filler1 TEXT,birth_date DATE, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS J_S( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,country_code TEXT,postal_code TEXT,city TEXT,c_o TEXT, filler2 TEXT,filler3 TEXT,filler4 TEXT,country_code_mail_address TEXT, postal_code_mail_address TEXT,city_mail_address TEXT,date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS T_AC( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,date_of_entry_into_group TEXT,date_of_leaving_group TEXT,reason_for_leaving_the_group TEXT,bank_code TEXT, bank_account_number TEXT,swift_bic TEXT,owner_of_bank_account TEXT,employee_with_potential DATE,indicator_for_executive TEXT,indicator_for_security_advisor TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS AD_AN( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,indicator_eligibility_stock_purchase_program TEXT,company_code TEXT,personal_number TEXT,plant_identifier_for_personal_number TEXT, filler5 TEXT,filler6 TEXT,transfer_date TEXT,filler7 TEXT,filler8 TEXT,local_cost_center TEXT, date_entry DATE DEFAULT today )")
    
    con.execute("CREATE TABLE IF NOT EXISTS AM_AW( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,employee_group TEXT,filler9 TEXT,department_abbreviation TEXT,hr_executive_level TEXT, employment_type TEXT,organizational_code TEXT,physical_work_location_code TEXT,internal_mail_code TEXT,level_of_business_allocation1 TEXT,level_of_business_allocation2 TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS AX_BG( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,mailing_language_of_employee TEXT,building TEXT,room_number TEXT,location_code TEXT, country_of_birth TEXT,city_of_birth TEXT,citizenship TEXT,filler10 TEXT,filler11 TEXT,street_and_house_number_of_home_address TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS BH_BQ( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,street_and_house_number_for_preferred_mailing_address TEXT,filler12 TEXT,job_code TEXT,plant2 TEXT, plant1 TEXT,organizational_code2 TEXT,smtp_email_address TEXT,position_entry_date TEXT,date_contract_end TEXT,hr_department TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS BR_CA( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,hr_representative TEXT,fax_number_hr_representative TEXT,management_level TEXT,work_phone_number_of_hr_representative TEXT, preferred_first_name TEXT,middle_initial TEXT,home_host_indicator TEXT,fulltime_parttime_equivalency TEXT,reports_to_kim TEXT,date_of_entry_into_company TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS CB_CW( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,filler12 TEXT,filler13 TEXT,filler14 TEXT, filler15 TEXT,filler16 TEXT,filler17 TEXT,filler18 TEXT,filler19 TEXT,filler20 TEXT,filler21 TEXT,filler22 TEXT, filler23 TEXT,filler24 TEXT,filler25 TEXT,filler26 TEXT,filler27 TEXT,filler28 TEXT,filler29 TEXT,filler30 TEXT,filler31 TEXT,filler32 TEXT,filler33 TEXT, date_entry DATE DEFAULT today )")                 

    con.execute("CREATE TABLE IF NOT EXISTS CX_DG( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,current_position_title TEXT,payment_type TEXT,plant_section TEXT,filler35 TEXT, indicator_for_disabled_person TEXT,social_security_number TEXT,state TEXT,state_for_preferred_mailing_address TEXT,phone_country_code TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS DH_DQ( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,phone_country_code TEXT,phone_extension_and_company TEXT,fax_country_code TEXT,fax_area_code TEXT, fax_extension_and_company TEXT,mobile_country_code TEXT,mobile_area_code TEXT,mobile_number TEXT,empl_id TEXT,filler36 TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS DR_EA( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,filler37 TEXT,bv_wechsel_ab TEXT,wohnhaft_bei TEXT,company_code_expat TEXT, currency TEXT,filler38 TEXT,year_of_goal_income TEXT,monthly_salary TEXT,delete_flag TEXT,filler39 TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS EB_EK( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,grade_band TEXT,shift_indicator TEXT,employee_category TEXT,department_name TEXT, status_employee TEXT,code_dl_tv TEXT,grade TEXT,nacos_cost_center TEXT,wage_group TEXT,year_of_wage_group TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS EL_EY( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,date_of_assignment_to_wage_group TEXT,filler40 TEXT,logon_id_of_employee TEXT,company_name_of_external_person TEXT, filler41 TEXT,filler42 TEXT,filler43 TEXT,filler44 TEXT,filler45 TEXT,filler46 TEXT,filler47 TEXT,filler48 TEXT,filler49 TEXT,filler50 TEXT, date_entry DATE DEFAULT today )")

    con.execute("CREATE TABLE IF NOT EXISTS EZ_FI( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,bv_wechsel_ab TEXT,filler51 TEXT,filler52 TEXT,type_of_timeframe_for_tax_data TEXT, starting_of_timeframe_for_tax_data TEXT,end_of_timeframe_for_tax_data TEXT,type_of_timeframe_for_tax_data2 TEXT, starting_of_timeframe_for_tax_data2 TEXT,end_of_timeframe_for_tax_data2 TEXT, type_of_timeframe_for_tax_data3 TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS FJ_FS( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,starting_of_timeframe_for_tax_data TEXT,end_of_timeframe_for_tax_data TEXT, type_of_timeframe_for_tax_data TEXT, starting_of_timeframe_for_tax_data2 TEXT,end_of_timeframe_for_tax_data2 TEXT, type_of_timeframe_for_tax_data2 TEXT ,starting_of_timeframe_for_tax_data3 TEXT,end_of_timeframe_for_tax_data3 TEXT, type_of_timeframe_for_tax_data3 TEXT,starting_of_timeframe_for_tax_data4 TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS FT_GC( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,end_of_timeframe_for_tax_data TEXT,type_of_timeframe_for_tax_data TEXT,starting_of_timeframe_for_tax_data TEXT,end_of_timeframe_for_tax_data2 TEXT,type_of_timeframe_for_tax_data2 TEXT,starting_of_timeframe_for_tax_data2 TEXT,end_of_timeframe_for_tax_data3 TEXT,type_of_timeframe_for_tax_data3 TEXT,starting_of_timeframe_for_tax_data3 TEXT,end_of_timeframe_for_tax_data4 TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS GD_GM( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,type_of_timeframe_for_tax_data TEXT,starting_of_timeframe_for_tax_data TEXT,end_of_timeframe_for_tax_data TEXT,type_of_timeframe_for_tax_data2 TEXT,starting_of_timeframe_for_tax_data2 TEXT,end_of_timeframe_for_tax_data2 TEXT,type_of_timeframe_for_tax_data3 TEXT,starting_of_timeframe_for_tax_data3 TEXT,end_of_timeframe_for_tax_data3 TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS GN_GW( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,permanent_residence TEXT,filler54 TEXT,wage_and_salary_group TEXT,diversity TEXT, executive_bonus TEXT,number_of_monthly_base_salary TEXT,structure_code TEXT,dept_id TEXT,full_time_equivalent TEXT,bonus_payout TEXT, date_entry DATE DEFAULT today )")
    con.execute("CREATE TABLE IF NOT EXISTS GX_HK( id INTEGER PRIMARY KEY AUTOINCREMENT ,kim TEXT,position_control_number TEXT,daimler_cost_center TEXT,last_name_passport TEXT,first_name_passport TEXT, executive_bonus TEXT,iban TEXT,local_job_profile TEXT,irwwh TEXT,buchungskreis TEXT,rabattkennzeichen TEXT,valid_by TEXT,widersprecherkennzeichen TEXT,ivv_contract TEXT,ivv_risk_taker TEXT,ivv_control_unit TEXT, date_entry DATE DEFAULT today )")
    con.close()
#-----------------------------------------------------------------------------------------------------------------------------------
def updateInfo(data):
        conn = sqlite3.connect('database.sqlite')
        con = conn.cursor()  
    
        sql =  ''' UPDATE A_I
                    SET first_name = ?,
                    last_name = ?,
                    gender = ?,
                    birth_date = ?
                    WHERE kim = ?'''
        con.execute(sql, (data.get('fname'),data.get('lname'),data.get('gender'),data.get('bdate'),data.get('kim')))

        sql =  '''UPDATE T_AC
        SET date_of_entry_into_group = ?
                    WHERE kim = ?'''
        con.execute(sql, [data.get('date_of_entry'),data.get('kim')])

        sql =  ''' UPDATE AD_AN
        SET plant_identifier_for_personal_number = ? 
                    WHERE kim = ?'''
        con.execute(sql, [data.get('pnumber'),data.get('kim')])

        sql =  ''' UPDATE AM_AW
        SET employee_group = ?,
                hr_executive_level = ?,
                employment_type = ?
                WHERE kim = ?'''
        con.execute(sql, [data.get('employee_group'),data.get('hr_level'),data.get('etype'),data.get('kim')])

        sql =  ''' UPDATE AM_AW
        SET hr_executive_level = ?
                    WHERE kim = ?'''
        con.execute(sql, [data.get('hr_level'),data.get('kim')])


        sql =  ''' UPDATE BH_BQ
        SET smtp_email_address = ?,
                date_contract_end=?
                    WHERE kim = ?'''
        con.execute(sql, [data.get('email'),data.get('date_of_contract_end'),data.get('kim')])

        sql =  ''' UPDATE BR_CA
        SET reports_to_kim = ?
                    WHERE kim = ?'''
        con.execute(sql, [data.get('report'),data.get('kim')])

        sql =  ''' UPDATE CX_DG
        SET current_position_title = ?
                    WHERE kim = ?'''
        con.execute(sql, [data.get('current_position'),data.get('kim')])
        conn.commit()
#-----------------------------------date update---------------------------------------
        sql =  ''' UPDATE A_I
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE J_S
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE T_AC
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE AD_AN
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE AM_AW
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE AX_BG
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE BH_BQ
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE BR_CA
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()


        sql =  ''' UPDATE CB_CW
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE CX_DG
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE EB_EK
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE DH_DQ
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE DR_EA
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE EL_EY
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE EZ_FI
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()


        sql =  ''' UPDATE FJ_FS
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE FT_GC
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE GD_GM
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE GN_GW
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE GX_HK
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()
        return 1
 #   except Exception as e:
  #      print(e)
  #      return 0
#-----------------------------------------------------------------------------------------------------------------------------------
def loadInfo():
    conn = sqlite3.connect('database.sqlite')
    con = conn.cursor()
    file = 'assets/test.xlsx'
    wb_obj = load_workbook(filename=file)
    ws = wb_obj.active
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            print('loading...')
            con.execute('''INSERT INTO A_I (kim,sin,gender,first_name,last_name,last_name_prefix,academic_title,filler1,birth_date,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value ,row[8].value,date.today() ))
            con.execute('''INSERT INTO J_S (kim,country_code,postal_code,city,c_o,filler2,filler3,filler4,country_code_mail_address,postal_code_mail_address,city_mail_address,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[9].value,row[10].value,row[11].value,row[12].value,row[13].value,row[14].value,row[15].value,row[16].value,row[17].value,row[18].value,date.today() ))
            con.execute('''INSERT INTO T_AC (kim,date_of_entry_into_group,date_of_leaving_group,reason_for_leaving_the_group,bank_code,bank_account_number,swift_bic,owner_of_bank_account,employee_with_potential,indicator_for_executive,indicator_for_security_advisor,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[19].value,row[20].value,row[21].value,row[22].value,row[23].value,row[24].value,row[25].value,row[26].value,row[27].value,row[28].value,date.today() ))
            con.execute('''INSERT INTO AD_AN (kim,indicator_eligibility_stock_purchase_program,company_code,personal_number,plant_identifier_for_personal_number,filler5,filler6,transfer_date,filler7,filler8,local_cost_center,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[29].value,row[30].value,row[31].value,row[32].value,row[33].value,row[34].value,row[35].value,row[36].value,row[37].value,row[38].value,date.today() ))
            con.execute('''INSERT INTO AM_AW (kim,employee_group,filler9,department_abbreviation,hr_executive_level,employment_type,organizational_code,physical_work_location_code,internal_mail_code,level_of_business_allocation1,level_of_business_allocation2,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[39].value,row[40].value,row[41].value,row[42].value,row[43].value,row[44].value,row[45].value,row[46].value,row[47].value,row[48].value,date.today() ))
            
            con.execute('''INSERT INTO AX_BG (kim,mailing_language_of_employee,building,room_number,location_code,country_of_birth,city_of_birth,citizenship,filler10,filler11,street_and_house_number_of_home_address,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[49].value,row[50].value,row[51].value,row[52].value,row[53].value,row[54].value,row[55].value,row[56].value,row[57].value,row[58].value,date.today() ))
            con.execute('''INSERT INTO BH_BQ (kim,street_and_house_number_for_preferred_mailing_address,filler12,job_code,plant1,plant2,organizational_code2,smtp_email_address,position_entry_date,date_contract_end,hr_department,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[59].value,row[60].value,row[61].value,row[62].value,row[64].value,row[63].value,row[65].value,row[66].value,row[67].value,row[68].value,date.today() ))
            con.execute('''INSERT INTO BR_CA (kim,hr_representative,fax_number_hr_representative,management_level,work_phone_number_of_hr_representative,preferred_first_name,middle_initial,home_host_indicator,fulltime_parttime_equivalency,reports_to_kim,date_of_entry_into_company,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[69].value,row[70].value,row[71].value,row[72].value,row[73].value,row[74].value,row[75].value,row[76].value,row[77].value,row[78].value,date.today() ))
            con.execute('''INSERT INTO CB_CW (kim,filler12,filler13,filler14,filler15,filler16,filler17,filler18,filler19,filler20,filler21,filler22,filler23,filler24,filler25,filler26,filler27,filler28,filler29,filler30,filler31,filler32,filler33,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[79].value,row[80].value,row[81].value,row[82].value,row[83].value,row[84].value,row[85].value,row[86].value,row[87].value,row[88].value,row[89].value,row[90].value,row[91].value,row[92].value,row[93].value,row[94].value,row[95].value,row[96].value,row[97].value,row[98].value,row[99].value,row[100].value,date.today() ))
            con.execute('''INSERT INTO CX_DG (kim,current_position_title,payment_type,plant_section,filler35,indicator_for_disabled_person,social_security_number,state,state_for_preferred_mailing_address,phone_country_code,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[101].value,row[98].value,row[102].value,row[103].value,row[104].value,row[105].value,row[106].value,row[107].value,row[108].value,date.today() ))
            
            con.execute('''INSERT INTO DH_DQ (kim,phone_country_code,phone_extension_and_company,fax_country_code,fax_area_code,fax_extension_and_company,mobile_country_code,mobile_area_code,mobile_number,empl_id,filler36,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[110].value,row[111].value,row[112].value,row[113].value,row[114].value,row[115].value,row[116].value,row[117].value,row[118].value,row[119].value,date.today() ))
            con.execute('''INSERT INTO DR_EA (kim,filler37,bv_wechsel_ab,wohnhaft_bei,company_code_expat,currency,filler38,year_of_goal_income,monthly_salary,delete_flag,filler39,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EB_EK (kim,grade_band,shift_indicator,employee_category,department_name,status_employee,status_employee,code_dl_tv,grade,nacos_cost_center,wage_group,year_of_wage_group,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EL_EY (kim,date_of_assignment_to_wage_group,filler40,logon_id_of_employee,company_name_of_external_person,filler41,filler42,filler43,filler44,filler45,filler46,filler47,filler48,filler49,filler50,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EZ_FI (kim,bv_wechsel_ab,filler51,filler52,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data3,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))

            con.execute('''INSERT INTO FJ_FS (kim,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data3,end_of_timeframe_for_tax_data3,type_of_timeframe_for_tax_data3,starting_of_timeframe_for_tax_data4,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO FT_GC (kim,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data3,type_of_timeframe_for_tax_data3,starting_of_timeframe_for_tax_data3,end_of_timeframe_for_tax_data4,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GD_GM (kim,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data3,starting_of_timeframe_for_tax_data3,end_of_timeframe_for_tax_data3,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GN_GW (kim,permanent_residence,wage_and_salary_group,diversity,executive_bonus,number_of_monthly_base_salary,structure_code,dept_id,full_time_equivalent,bonus_payout,filler54,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[203].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GX_HK (kim,executive_bonus,rabattkennzeichen,position_control_number,daimler_cost_center,last_name_passport,last_name_passport,last_name_passport,first_name_passport,iban,local_job_profile,irwwh,valid_by,buchungskreis,widersprecherkennzeichen,ivv_contract,ivv_risk_taker,ivv_control_unit,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            conn.commit()
#-----------------------------------------------------------------------------------------------------------------------------------
def saveInfo(data):
    conn = sqlite3.connect('database.sqlite')
    con = conn.cursor()  
    file = 'assets/test.xlsx'
    wb_obj = load_workbook(filename=file)
    ws = wb_obj.active
    for row in ws.iter_rows(min_row=10, min_col=1, max_row=11, max_col=ws.max_column):
            con.execute('''INSERT INTO A_I (kim,sin,gender,first_name,last_name,last_name_prefix,academic_title,filler1,birth_date,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[1].value,data.get('gender'),data.get('fname'),data.get('lname'),row[5].value,row[6].value,row[7].value, data.get('bdate'),date.today() ))
            con.execute('''INSERT INTO J_S (kim,country_code,postal_code,city,c_o,filler2,filler3,filler4,country_code_mail_address,postal_code_mail_address,city_mail_address,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[9].value,row[10].value,row[11].value,row[12].value,row[13].value,row[14].value,row[15].value,row[16].value,row[17].value,row[18].value,date.today() ))
            con.execute('''INSERT INTO T_AC (kim,date_of_entry_into_group,date_of_leaving_group,reason_for_leaving_the_group,bank_code,bank_account_number,swift_bic,owner_of_bank_account,employee_with_potential,indicator_for_executive,indicator_for_security_advisor,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),data.get('date_of_entry'),row[20].value,row[21].value,row[22].value,row[23].value,row[24].value,row[25].value,row[26].value,row[27].value,row[28].value,date.today() ))
            con.execute('''INSERT INTO AD_AN (kim,indicator_eligibility_stock_purchase_program,company_code,personal_number,plant_identifier_for_personal_number,filler5,filler6,transfer_date,filler7,filler8,local_cost_center,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[29].value,row[30].value,row[31].value,data.get('personal_number'),row[33].value,row[34].value,row[35].value,row[36].value,row[37].value,row[38].value,date.today() ))
            con.execute('''INSERT INTO AM_AW (kim,employee_group,filler9,department_abbreviation,hr_executive_level,employment_type,organizational_code,physical_work_location_code,internal_mail_code,level_of_business_allocation1,level_of_business_allocation2,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),data.get('employee_group'),row[40].value,row[41].value,data.get('hr_level'),data.get('etype'),row[44].value,row[45].value,row[46].value,row[47].value,row[48].value,date.today() ))
            con.execute('''INSERT INTO AX_BG (kim,mailing_language_of_employee,building,room_number,location_code,country_of_birth,city_of_birth,citizenship,filler10,filler11,street_and_house_number_of_home_address,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[49].value,row[50].value,row[51].value,row[52].value,row[53].value,row[54].value,data.get('nationality'),row[56].value,row[57].value,row[58].value,date.today() ))
            con.execute('''INSERT INTO BH_BQ (kim,street_and_house_number_for_preferred_mailing_address,filler12,job_code,plant1,plant2,organizational_code2,smtp_email_address,position_entry_date,date_contract_end,hr_department,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[59].value,row[60].value,row[61].value,row[62].value,row[64].value,row[63].value,data.get('temail'),data.get('date_of_entry'),data.get('date_of_contract_end'),row[68].value,date.today() ))
            con.execute('''INSERT INTO BR_CA (kim,hr_representative,fax_number_hr_representative,management_level,work_phone_number_of_hr_representative,preferred_first_name,middle_initial,home_host_indicator,fulltime_parttime_equivalency,reports_to_kim,date_of_entry_into_company,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[69].value,row[70].value,row[71].value,row[72].value,row[73].value,row[74].value,row[75].value,row[76].value,data.get('report'),data.get('date_of_entry'),date.today() ))
            con.execute('''INSERT INTO CB_CW (kim,filler12,filler13,filler14,filler15,filler16,filler17,filler18,filler19,filler20,filler21,filler22,filler23,filler24,filler25,filler26,filler27,filler28,filler29,filler30,filler31,filler32,filler33,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[79].value,row[80].value,row[81].value,row[82].value,row[83].value,row[84].value,row[85].value,row[86].value,row[87].value,row[88].value,row[89].value,row[90].value,row[91].value,row[92].value,row[93].value,row[94].value,row[95].value,row[96].value,row[97].value,row[98].value,row[99].value,row[100].value,date.today() ))
            con.execute('''INSERT INTO CX_DG (kim,current_position_title,payment_type,plant_section,filler35,indicator_for_disabled_person,social_security_number,state,state_for_preferred_mailing_address,phone_country_code,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),data.get('position'),row[98].value,row[102].value,row[103].value,row[104].value,row[105].value,row[106].value,row[107].value,row[108].value,date.today() ))
            
            con.execute('''INSERT INTO DH_DQ (kim,phone_country_code,phone_extension_and_company,fax_country_code,fax_area_code,fax_extension_and_company,mobile_country_code,mobile_area_code,mobile_number,empl_id,filler36,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[110].value,row[111].value,row[112].value,row[113].value,row[114].value,row[115].value,row[116].value,row[117].value,row[118].value,row[119].value,date.today() ))
            con.execute('''INSERT INTO DR_EA (kim,filler37,bv_wechsel_ab,wohnhaft_bei,company_code_expat,currency,filler38,year_of_goal_income,monthly_salary,delete_flag,filler39,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EB_EK (kim,grade_band,shift_indicator,employee_category,department_name,status_employee,status_employee,code_dl_tv,grade,nacos_cost_center,wage_group,year_of_wage_group,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EL_EY (kim,date_of_assignment_to_wage_group,filler40,logon_id_of_employee,company_name_of_external_person,filler41,filler42,filler43,filler44,filler45,filler46,filler47,filler48,filler49,filler50,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO EZ_FI (kim,bv_wechsel_ab,filler51,filler52,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data3,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
           
            con.execute('''INSERT INTO FJ_FS (kim,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data3,end_of_timeframe_for_tax_data3,type_of_timeframe_for_tax_data3,starting_of_timeframe_for_tax_data4,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO FT_GC (kim,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data3,type_of_timeframe_for_tax_data3,starting_of_timeframe_for_tax_data3,end_of_timeframe_for_tax_data4,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(row[0].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GD_GM (kim,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data3,starting_of_timeframe_for_tax_data3,end_of_timeframe_for_tax_data3,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GN_GW (kim,permanent_residence,wage_and_salary_group,diversity,executive_bonus,number_of_monthly_base_salary,structure_code,dept_id,full_time_equivalent,bonus_payout,filler54,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,1,row[120].value,row[120].value,date.today() ))
            con.execute('''INSERT INTO GX_HK (kim,executive_bonus,rabattkennzeichen,position_control_number,daimler_cost_center,last_name_passport,last_name_passport,last_name_passport,first_name_passport,iban,local_job_profile,irwwh,valid_by,buchungskreis,widersprecherkennzeichen,ivv_contract,ivv_risk_taker,ivv_control_unit,date_entry)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ''',(data.get("kim"),row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,row[120].value,date.today() ))
            conn.commit()
            return 1
#---------------------------------------------------------------------------------------------------------------------
def check_kim(size):
    return size == 15
#---------------------------------------------------------------------------------------------------------------------
def gen2(title):
    file = 'assets/report2.xlsx'
    wb_obj = load_workbook(filename=file)
    ws = wb_obj.active
    ws["d7"] = "one"
    wb_obj.save(filename="Reports2/"+title+".xlsx")
    return 1
#---------------------------------------------------------------------------------------------------------------------
def find_employee(kim):
    conn = sqlite3.connect('database.sqlite')
    con = conn.cursor()
    con.execute("SELECT * FROM A_I WHERE kim=?", (str(kim),))
    rows = con.fetchall()
    if len(rows) > 0:
        return 1
    else:
        return 0
#---------------------------------------------------------------------------------------------------------------------
def collect_data(kim):
    data = []
   # kim='123456789012345'
    conn = sqlite3.connect('database.sqlite')
    con = conn.cursor()
    con.execute("SELECT gender,first_name,last_name,birth_date  FROM A_I Where kim = ?",[kim])
    data.extend(list(con.fetchall()))
    con.execute("SELECT date_of_entry_into_group FROM T_AC Where kim= ?",[kim])
    data.extend(list(con.fetchall()))
    con.execute("SELECT plant_identifier_for_personal_number FROM AD_AN Where kim= ?",[kim])
    data.extend(list(con.fetchall()))
    con.execute("SELECT employee_group,hr_executive_level,employment_type FROM AM_AW Where kim= ?",[kim])
    data.extend(list(con.fetchall()))
    con.execute("SELECT smtp_email_address,date_contract_end FROM BH_BQ Where kim =?",[kim])
    data.extend(list(con.fetchall()))
    con.execute("SELECT reports_to_kim FROM BR_CA Where kim =?",[kim])
    data.extend(list(con.fetchall()))
    con.execute("SELECT current_position_title FROM CX_DG Where kim= ?",[kim])
    data.extend(list(con.fetchall()))
    out = None
    out = {'gender':data[0][0],'name':str(data[0][1]).strip(),'surname':str(data[0][2]).strip(),'birthday':fix_d(data[0][3]),'date':data[1][0],'pnum':data[2][0],'eg':data[3][0],'hr':data[3][1],'et':data[3][2],'email':str(data[4][0]).strip(),'contract_end':str(data[4][1]).strip(),'report':data[5][0],'position':str(data[6][0]).strip()}
    return out

#---------------------------------------------------------------------------------------------------------------------
def terminate(data):
        conn = sqlite3.connect('database.sqlite')
        con = conn.cursor()
        sql =  ''' UPDATE T_AC
        SET date_of_leaving_group = ?,
                reason_for_leaving_the_group = ?,
                date_entry = ?
                WHERE kim = ?'''
        print(data)
        con.execute(sql, [data.get('date'),data.get('reason'),date.today(),data.get('kim')])
#-----------------------------------date update---------------------------------------
        sql =  ''' UPDATE A_I
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE J_S
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE T_AC
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE AD_AN
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE AM_AW
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE AX_BG
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE BH_BQ
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE BR_CA
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()


        sql =  ''' UPDATE CB_CW
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE CX_DG
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE EB_EK
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE DH_DQ
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE DR_EA
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE EL_EY
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE EZ_FI
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()


        sql =  ''' UPDATE FJ_FS
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE FT_GC
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE GD_GM
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE GN_GW
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()

        sql =  ''' UPDATE GX_HK
        SET date_entry = ?
                    WHERE kim = ?'''
        con.execute(sql, [date.today(),data.get('kim')])
        conn.commit()
        return 1

#---------------------------------------------------------------------------------------------------------------------
def collect_data2(kim):
    data = []
    conn = sqlite3.connect('database.sqlite')
    con = conn.cursor()
    try:
        con.execute("SELECT first_name,last_name  FROM A_I Where kim = ?",[kim])
        data.extend(list(con.fetchall()))
        return {'name':str(data[0][0]).strip(),'surname':str(data[0][1]).strip()}
    except Exception as e:
        print(e)
        return None
#---------------------------------------------------------------------------------------------------------------------
def generate_report(start,stop):
    data = []
    conn = sqlite3.connect('database.sqlite')
    con = conn.cursor()
    con.execute("SELECT kim,sin,gender,first_name,last_name,last_name_prefix,academic_title,filler1,birth_date  FROM A_I Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows1 = con.fetchall()
    con.execute("SELECT country_code,postal_code,city,c_o,filler2,filler3,filler4,country_code_mail_address,postal_code_mail_address,city_mail_address FROM J_S Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows2 = con.fetchall()
    con.execute("SELECT date_of_entry_into_group,date_of_leaving_group,reason_for_leaving_the_group, bank_code,bank_account_number,swift_bic,owner_of_bank_account,employee_with_potential,indicator_for_executive,indicator_for_security_advisor FROM T_AC Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows3 = con.fetchall()
    con.execute("SELECT indicator_eligibility_stock_purchase_program,company_code,personal_number,plant_identifier_for_personal_number,filler5,filler6,transfer_date,filler7,filler8,local_cost_center FROM AD_AN Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows4 = con.fetchall()
    con.execute("SELECT employee_group,filler9,department_abbreviation,hr_executive_level,employment_type,organizational_code,physical_work_location_code,internal_mail_code,level_of_business_allocation1,level_of_business_allocation2 FROM AM_AW Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows5 = con.fetchall()
    con.execute("SELECT mailing_language_of_employee,building,room_number,location_code,country_of_birth,city_of_birth,citizenship,filler10,filler11,street_and_house_number_of_home_address FROM AX_BG Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows6 = con.fetchall()
    con.execute("SELECT street_and_house_number_for_preferred_mailing_address,filler12,job_code,plant2,plant1,organizational_code2,smtp_email_address,position_entry_date,date_contract_end,hr_department FROM BH_BQ Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows7 = con.fetchall()
    con.execute("SELECT hr_representative,fax_number_hr_representative,management_level,work_phone_number_of_hr_representative,preferred_first_name,middle_initial,home_host_indicator,fulltime_parttime_equivalency,reports_to_kim,date_of_entry_into_company FROM BR_CA Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows8 = con.fetchall()
    con.execute("SELECT filler12,filler13,filler14,filler15,filler16,filler17,filler18,filler19,filler20,filler21,filler22,filler23,filler24,filler25,filler26,filler27,filler28,filler29,filler30,filler31,filler32,filler33 FROM CB_CW Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows9 = con.fetchall()
    con.execute("SELECT current_position_title,payment_type,plant_section,filler35,indicator_for_disabled_person,social_security_number,state,state_for_preferred_mailing_address,phone_country_code FROM CX_DG Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows10 = con.fetchall()
    con.execute("SELECT phone_country_code,phone_extension_and_company,fax_country_code,fax_area_code,fax_extension_and_company,mobile_country_code,mobile_area_code,mobile_number,empl_id,filler36 FROM DH_DQ Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows11 = con.fetchall()
    con.execute("SELECT filler37,bv_wechsel_ab,wohnhaft_bei,company_code_expat,currency,filler38,year_of_goal_income,monthly_salary,delete_flag,filler39 FROM DR_EA Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows12 = con.fetchall()
    con.execute("SELECT grade_band,shift_indicator,employee_category,department_name,status_employee,code_dl_tv,grade,nacos_cost_center,wage_group,year_of_wage_group FROM EB_EK Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows13 = con.fetchall()
    con.execute("SELECT date_of_assignment_to_wage_group,filler40,logon_id_of_employee,company_name_of_external_person,filler41,filler42,filler43,filler44,filler45,filler46,filler47,filler48,filler49,filler50 FROM EL_EY Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows14 = con.fetchall()
    con.execute("SELECT bv_wechsel_ab,filler51,filler52,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data3 FROM EZ_FI Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows15 = con.fetchall()
    con.execute("SELECT starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data3,end_of_timeframe_for_tax_data3,type_of_timeframe_for_tax_data3,starting_of_timeframe_for_tax_data4 FROM FJ_FS Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows16 = con.fetchall()
    con.execute("SELECT end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data3,type_of_timeframe_for_tax_data3,starting_of_timeframe_for_tax_data3,end_of_timeframe_for_tax_data4 FROM FT_GC Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows18= con.fetchall()
    con.execute("SELECT type_of_timeframe_for_tax_data,starting_of_timeframe_for_tax_data,end_of_timeframe_for_tax_data,type_of_timeframe_for_tax_data2,starting_of_timeframe_for_tax_data2,end_of_timeframe_for_tax_data2,type_of_timeframe_for_tax_data3,starting_of_timeframe_for_tax_data3,end_of_timeframe_for_tax_data3 FROM GD_GM Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows19 = con.fetchall()
    con.execute("SELECT permanent_residence,filler54,wage_and_salary_group,diversity,executive_bonus,number_of_monthly_base_salary,structure_code,dept_id,full_time_equivalent,bonus_payout FROM GN_GW Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows20 = con.fetchall()
    con.execute("SELECT position_control_number,daimler_cost_center,last_name_passport,first_name_passport,executive_bonus,iban,local_job_profile,irwwh,buchungskreis,rabattkennzeichen,valid_by,widersprecherkennzeichen,ivv_contract,ivv_risk_taker,ivv_control_unit FROM GX_HK Where date_entry BETWEEN date(?) and date(?)",[start,stop])
    rows21 = con.fetchall()
    
    out = []
    for x in range(0,len(rows2)):
        data.extend(list(rows1[x]))
        data.extend(list(rows2[x]))
        data.extend(list(rows3[x]))
        data.extend(list(rows4[x]))
        data.extend(list(rows5[x]))
        data.extend(list(rows6[x]))
        data.extend(list(rows7[x]))
        data.extend(list(rows8[x]))
        data.extend(list(rows9[x]))
        data.extend(list(rows10[x]))
        data.extend(list(rows11[x]))
        data.extend(list(rows12[x]))
        data.extend(list(rows13[x]))
        data.extend(list(rows14[x]))
        data.extend(list(rows15[x]))
        data.extend(list(rows16[x]))
        data.extend(list(rows18[x]))
        data.extend(list(rows19[x]))
        data.extend(list(rows20[x]))
        data.extend(list(rows21[x]))
        out.append(data)
        data = []
    return [out,len(rows1)]
#-----------------------------------------------------------------------------------------------------------------------------------
def process(wording, spacing):
    max_len = len(wording)
    padding = abs(spacing-max_len)
    for i in range(0, padding):
        wording = wording+" "
    return wording
#-----------------------------------------------------------------------------------------------------------------------------------
def process1(wording, spacing):
    max_len = len(wording)
    padding = abs(168-max_len)
    for i in range(0, padding):
        wording = wording+" "
    return wording
#-----------------------------------------------------------------------------------------------------------------------------------
def gen(data, title):
	count =0
	if data[1] > 0:
		with open('Reports/'+title+'.txt','w', encoding = 'utf-8') as f:
			f.write('HEADERFIX      MDAT589   02112021021120211              WALTER MUBAI                  WALTER.MUBAI@ATLANTISFOUNDRIES.COM                0027215737382')
			for i in data[0]:
				f.write('\n')
				count =0
				for j in i:
					count = count + 1
					if count == 66 and j[0] != ' ':
						f.writelines(process1(j,12))
					elif count == 9:
						f.writelines(process(fix_d(j),12))
					else:
						f.writelines((process(str(j),12).replace('.','')).replace('-',''))
		
		return count
	else:
		return 0
#---------------------------------------------------------------------------------------------------------------------
def dateCheck(arg,value):
            result = []
            maps=[]
            
            if len(value[0])==1:
                maps.append(1)
            else:
                maps.append(0)
            if len(value[1])== 1:
                maps.append(1)
            else:
                maps.append(0)
            if arg == 'to':
                if maps[0] ==1:
                       result.append('-0')
                else:
                        result.append('-')
                if maps[1] ==1:
                       result.append('-0')
                else:
                        result.append('-')
                return value[2]+result[1]+value[1]+result[0]+value[0]
            else:
                if maps[0] ==1:
                       result.append('-0')
                else:
                        result.append('-')
                if maps[1] ==1:
                       result.append('-0')
                else:
                        result.append('-')
               
                return value[2]+result[1]+value[1]+result[0]+value[0]
#---------------------------------------------------------------------------------------------------------------------
def dateCheck1(value):
        value = str(value).split('-')
        return value[2]+value[1]+value[0]
#---------------------------------------------------------------------------------------------------------------------
def fix_input(option, word):
    if option == 'fn' or option == 'ln':
        mx = 30
        wlen = len(word)
        for x in range(0 ,(mx-wlen)):
            word = word+' '
        return word

    elif option == 'kim' or option == 'rk':
        mx = 15
        wlen = len(word)
        for x in range(0 ,(mx-wlen)):
            word = word+' '
        return word

    elif option == 'pn':
        mx = 4
        wlen = len(word)
        for x in range(0 ,(mx-wlen)):
            word = word+' '
        return word

    elif option == 'pos':
        mx = 65
        wlen = len(word)
        for x in range(0 ,(mx-wlen)):
            word = word+' '
        return word

    elif option == 'eg':
        return word

    elif option == 'nat':
        return word

    elif option == 'hr':
        return word

    elif option == 'et':
        return word

    elif option == 'email':
        mx = 90
        wlen = len(word)
        for x in range(mx-wlen):
            word = word+' '
        return word
#------------------------------------------------------------------------------------------------------
def fix_value(word):
    mx_len = len(word)
    padding = 9- mx_len
    for x in range(padding):
        word = '0'+ word
    return word
#------------------------------------------------------------------------------------------------------
def fix_d(word):
    if len(str(word)) != 8:
        return '0'+ str(word)
    return str(word)
#------------------------------------------------------------------------------------------------------
def comboFix(field, option):
    if field == 'gender':
        if option == 'Male':
            return 1
        else:
            return 2
    elif field == 'area':
        if option == 'Support':
            return 1
        elif option == 'Direct':
            return 3
        elif option == 'Indirect':
            return 4
    elif field == 'job':
        return option
    elif field == 'nationality':
        return option
    elif field == 'contract':
        if option == 'Permanent':
            return 100
        elif  option =='LTD Duration':
            return 210
        elif option == 'Apprentice Ext':
            return 311
        elif option == 'Students':
            return 330
    elif field == 'reason':
        if option == 'Death':
            return 21
        elif option == 'Retirement':
            return 22
        elif option =='Mutual Agreement':
            return 24
        elif option=='Dismissal/Termination':
            return 25
        elif option == 'Selling of Company':
            return 26
        elif option == 'Self-Motivated Termination of Contract':
            return 27
        elif option == 'Company Driven Termination':
            return 28
        elif option == 'SA with Reemployment Promise (Layoffs)':
            return 29
#------------------------------------------------------------------------------------------------------
def fixCombo(option):
    option = int(option)
    if option== 1:
        return 'Male'
    elif option==2:
        return 'Female'
    elif option==29:
        return 'SA with Reemployment Promise (Layoffs)'
    elif option==28:
        return 'Company Driven Termination'
    elif option==27:
        return 'Self-Motivated Termination of Contract'
    elif option==26:
        return 'Selling of Company'
    elif option==25:
        return 'Dismissal/Termination'
    elif option==24:
        return 'Mutual Agreement'
    elif option==22:
        return 'Retirement'
    elif option==21:
        return 'Death'
    elif option==330:
        return 'Students'
    elif option==311:
        return 'Apprentice Ext'
    elif option==210:
        return 'LTD Duration'
    elif option==100:
        return 'Permanent'
    elif option==4:
        return 'Indirect'
    elif option==3:
        return 'Direct'
    elif option==1:
        return 'Support'
    
#------------------------------------------------------------------------------------------------------


